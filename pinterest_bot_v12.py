"""
Pinterest Niche Research Bot  v6 (Firefox)
=================================
FIXES in v3:
  - Uses Firefox + geckodriver (works perfectly on Python 3.14)
  - Auto-installs all required packages
  - Opens Pinterest homepage first before doing anything
  - Handles login popup automatically

HOW TO RUN (Windows CMD):
  py pinterest_bot_v6.py
"""

import sys
import subprocess

# ── Auto-install all required packages ──
def install_pkg(pkg):
    print(f"[*] Installing {pkg}...")
    subprocess.check_call([sys.executable, "-m", "pip", "install", pkg, "-q"])

packages = {
    
    "selenium":                "selenium",
    "openpyxl":                "openpyxl",
}
for module, pip_name in packages.items():
    try:
        __import__(module)
    except ImportError:
        install_pkg(pip_name)

# ── Imports ──
import time
import re
import os
from datetime import datetime

from selenium import webdriver
from selenium.webdriver.firefox.options import Options as FirefoxOptions
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException, NoSuchElementException, WebDriverException
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


# ─────────────────────────────────────────────
#  CONFIG
# ─────────────────────────────────────────────
MIN_PINTEREST_VIEWS_M = 2      # Minimum Pinterest monthly views (millions)
MIN_WEBSITE_TRAFFIC   = 200    # Minimum monthly website visitors
MAX_PINS_PER_SEARCH   = 20     # Only first 20 pins per search
SCROLL_PAUSE          = 2.5
PAGE_LOAD_WAIT        = 6
AUTOCOMPLETE_WAIT     = 3
AHREFS_WAIT           = 7

# ─────────────────────────────────────────────
#  PROXY LIST (Webshare.io)
#  Format: ip:port:username:password
# ─────────────────────────────────────────────
PROXIES = [
    "31.59.20.176:6754:lqsgsfmg:cup5c6a0szki",
    "198.23.239.134:6540:lqsgsfmg:cup5c6a0szki",
    "45.38.107.97:6014:lqsgsfmg:cup5c6a0szki",
    "107.172.163.27:6543:lqsgsfmg:cup5c6a0szki",
    "198.105.121.200:6462:lqsgsfmg:cup5c6a0szki",
    "216.10.27.159:6837:lqsgsfmg:cup5c6a0szki",
    "142.111.67.146:5611:lqsgsfmg:cup5c6a0szki",
    "191.96.254.138:6185:lqsgsfmg:cup5c6a0szki",
    "31.58.9.4:6077:lqsgsfmg:cup5c6a0szki",
    "104.239.107.47:5699:lqsgsfmg:cup5c6a0szki",
]

import random

def get_random_proxy():
    """Pick a random proxy from the list."""
    proxy = random.choice(PROXIES)
    parts = proxy.split(":")
    ip, port, username, password = parts[0], parts[1], parts[2], parts[3]
    return ip, port, username, password


# ─────────────────────────────────────────────
#  SETUP DRIVER  (Firefox — works on Python 3.14)
# ─────────────────────────────────────────────
def get_proxy():
    """
    Get proxy from environment variable.
    Format: PROXY_URL=http://username:password@proxy_host:port
    Set this in Railway environment variables.
    Get proxies from webshare.io
    """
    import os
    proxy_url = os.environ.get("PROXY_URL", "")
    if proxy_url:
        print(f"[+] Using proxy: {proxy_url.split('@')[-1]}")  # hide credentials
        return proxy_url
    print("[!] No proxy set — bot may be blocked by Pinterest on cloud")
    return None


def setup_driver():
    import os
    import geckodriver_autoinstaller
    geckodriver_autoinstaller.install()

    options = FirefoxOptions()

    # Detect if running on Railway/cloud server
    is_cloud = os.environ.get("RAILWAY_ENVIRONMENT") or os.environ.get("RAILWAY_SERVICE_NAME")

    if is_cloud:
        # Headless mode for Railway server
        print("[*] Running on Railway — headless Firefox mode")
        options.add_argument("--headless")
        options.add_argument("--no-sandbox")
        options.add_argument("--disable-dev-shm-usage")
        options.add_argument("--width=1280")
        options.add_argument("--height=1024")

        # ── Add proxy ──
        ip, port, username, password = get_random_proxy()
        print(f"[*] Using proxy: {ip}:{port}")

        # Set proxy via Firefox preferences
        options.set_preference("network.proxy.type", 1)
        options.set_preference("network.proxy.http", ip)
        options.set_preference("network.proxy.http_port", int(port))
        options.set_preference("network.proxy.ssl", ip)
        options.set_preference("network.proxy.ssl_port", int(port))
        options.set_preference("network.proxy.socks", ip)
        options.set_preference("network.proxy.socks_port", int(port))
        options.set_preference("network.proxy.socks_version", 5)
        options.set_preference("network.proxy.no_proxies_on", "localhost,127.0.0.1")

        # Proxy authentication via env
        os.environ["http_proxy"]  = f"http://{username}:{password}@{ip}:{port}"
        os.environ["https_proxy"] = f"http://{username}:{password}@{ip}:{port}"
    else:
        # Local mode — use real Firefox profile
        print("[*] Launching Firefox locally — please wait...")
        print("[!] Make sure Firefox is fully closed before continuing!")
        import shutil
        real_profile = r"C:\Users\Zarafshan\AppData\Roaming\Mozilla\Firefox\Profiles\7541cgod.default-release"
        temp_profile  = r"C:\Users\Zarafshan\AppData\Local\Temp\pinterest_bot_profile"
        if os.path.exists(temp_profile):
            shutil.rmtree(temp_profile, ignore_errors=True)
        try:
            shutil.copytree(real_profile, temp_profile)
            print("[+] Firefox profile copied successfully")
            options.add_argument("-profile")
            options.add_argument(temp_profile)
        except Exception as e:
            print(f"[!] Could not copy profile ({e}) — using fresh profile")

    # ── Add proxy if available ──
    proxy_url = get_proxy()
    if proxy_url:
        try:
            from urllib.parse import urlparse
            parsed = urlparse(proxy_url)
            proxy_host = parsed.hostname
            proxy_port = parsed.port
            proxy_user = parsed.username
            proxy_pass = parsed.password

            from selenium.webdriver.common.proxy import Proxy, ProxyType
            proxy = Proxy()
            proxy.proxy_type = ProxyType.MANUAL
            proxy.http_proxy = f"{proxy_host}:{proxy_port}"
            proxy.ssl_proxy  = f"{proxy_host}:{proxy_port}"
            options.proxy = proxy

            # Set auth if username/password provided
            if proxy_user and proxy_pass:
                options.set_preference("network.proxy.type", 1)
                options.set_preference("network.proxy.http", proxy_host)
                options.set_preference("network.proxy.http_port", proxy_port)
                options.set_preference("network.proxy.ssl", proxy_host)
                options.set_preference("network.proxy.ssl_port", proxy_port)
                options.set_preference("network.proxy.no_proxies_on", "")

            print(f"[+] Proxy configured: {proxy_host}:{proxy_port}")
        except Exception as e:
            print(f"[!] Proxy setup error: {e} — continuing without proxy")

    from selenium.webdriver.firefox.service import Service as FirefoxService
    driver = webdriver.Firefox(options=options)
    if not is_cloud:
        driver.maximize_window()
    driver.set_page_load_timeout(30)
    driver.set_script_timeout(30)

    # Handle proxy auth popup if it appears
    if proxy_url and is_cloud:
        try:
            from urllib.parse import urlparse
            parsed = urlparse(proxy_url)
            if parsed.username and parsed.password:
                driver.execute_script(f"""
                    window.addEventListener('load', function() {{
                        if (document.querySelector('[type=password]')) {{
                            document.querySelectorAll('input')[0].value = '{parsed.username}';
                            document.querySelectorAll('input')[1].value = '{parsed.password}';
                            document.querySelector('button').click();
                        }}
                    }});
                """)
        except Exception:
            pass

    return driver


# ─────────────────────────────────────────────
#  HELPERS
# ─────────────────────────────────────────────
def parse_number(text):
    if not text:
        return None
    t = text.strip().upper().replace(",", "").replace("+", "").replace(" ", "")
    try:
        if "B" in t:
            return int(float(t.replace("B", "")) * 1_000_000_000)
        if "M" in t:
            return int(float(t.replace("M", "")) * 1_000_000)
        if "K" in t:
            return int(float(t.replace("K", "")) * 1_000)
        return int(float(re.sub(r"[^\d.]", "", t)))
    except Exception:
        return None

def fmt(n):
    if n is None:
        return "N/A"
    if n >= 1_000_000:
        return f"{n/1_000_000:.1f}M"
    if n >= 1_000:
        return f"{n/1_000:.1f}K"
    return str(n)

def dismiss_popup(driver):
    for sel in [
        "button[data-test-id='closeup-close-button']",
        "button[aria-label='Close']",
        "div[data-test-id='interstitial-close-button']",
        "[data-test-id='login-page'] button[aria-label='Close']",
    ]:
        try:
            btn = driver.find_element(By.CSS_SELECTOR, sel)
            btn.click()
            time.sleep(0.8)
            return
        except Exception:
            continue


# ─────────────────────────────────────────────
#  OPEN PINTEREST & HANDLE LOGIN
# ─────────────────────────────────────────────
def open_pinterest(driver):
    print("[*] Opening Pinterest...")
    driver.get("https://www.pinterest.com/")
    time.sleep(PAGE_LOAD_WAIT)
    dismiss_popup(driver)

    # Check if login wall appeared
    try:
        current = driver.current_url
        page_src = driver.page_source.lower()
        if ("login" in current or
                "accounts" in current or
                "sign up" in page_src or
                "log in" in page_src[:2000]):
            print("\n" + "!"*55)
            print("  Pinterest wants you to log in.")
            print("  > Log in manually in the Firefox window.")
            print("  > Come back here and press Enter when done.")
            print("!"*55)
            input("  Press Enter after logging in: ")
            time.sleep(3)
            dismiss_popup(driver)
            print("[+] Continuing...")
        else:
            print("[+] Pinterest loaded. Bot is starting...")
    except Exception:
        pass


# ─────────────────────────────────────────────
#  GET SUB-NICHES FROM AUTOCOMPLETE
# ─────────────────────────────────────────────
def get_sub_niches(driver, main_niche):
    print(f"\n[*] Finding sub-niches for '{main_niche}'...")
    sub_niches = []

    try:
        driver.get("https://www.pinterest.com/")
        time.sleep(PAGE_LOAD_WAIT)
        dismiss_popup(driver)

        # Find search box
        search_box = None
        for sel in [
            "input[data-test-id='search-box-input']",
            "input[name='searchBoxInput']",
            "input[placeholder*='Search']",
            "input[type='text']",
        ]:
            try:
                search_box = WebDriverWait(driver, 8).until(
                    EC.presence_of_element_located((By.CSS_SELECTOR, sel))
                )
                if search_box:
                    break
            except Exception:
                continue

        if not search_box:
            raise Exception("Search box not found")

        search_box.clear()
        search_box.send_keys(main_niche)
        time.sleep(AUTOCOMPLETE_WAIT)

        # Grab suggestions
        for sel in [
            "[data-test-id='typeahead-term-item']",
            "ul[role='listbox'] li",
            "[class*='typeahead'] li",
            "[class*='Typeahead'] [role='option']",
            "[class*='suggestion']",
        ]:
            try:
                items = driver.find_elements(By.CSS_SELECTOR, sel)
                if items:
                    for item in items[:10]:
                        text = item.text.strip().lower()
                        if text and text != main_niche.lower() and len(text) > 3:
                            sub_niches.append(text)
                    if sub_niches:
                        break
            except Exception:
                continue

    except Exception as e:
        print(f"[!] Autocomplete error: {e}")

    # Fallback
    if not sub_niches:
        print("[!] Using pattern fallback for sub-niches")
        patterns = ["ideas", "aesthetic", "diy", "modern", "small",
                    "living room", "bedroom", "cheap", "inspiration", "apartment"]
        sub_niches = [f"{main_niche} {p}" for p in patterns]

    sub_niches = list(dict.fromkeys([s.strip() for s in sub_niches if s]))[:8]
    print(f"[+] Sub-niches: {sub_niches}")
    return sub_niches


# ─────────────────────────────────────────────
#  GET WEBSITE URL FROM PINTEREST PROFILE
# ─────────────────────────────────────────────
def get_website_from_profile(driver, profile_url):
    try:
        driver.get(profile_url)
        time.sleep(PAGE_LOAD_WAIT)
        dismiss_popup(driver)

        for sel in [
            "a[data-test-id='profile-website']",
            "a[data-test-id='user-website-url']",
            "[data-test-id='profile-website-link'] a",
            "[class*='websiteUrl'] a",
            "[class*='website'] a",
        ]:
            try:
                el = driver.find_element(By.CSS_SELECTOR, sel)
                href = el.get_attribute("href")
                if href and "pinterest.com" not in href and href.startswith("http"):
                    return href.strip()
            except Exception:
                continue

        # Fallback: scan all external links
        all_links = driver.find_elements(By.CSS_SELECTOR, "a[href]")
        for link in all_links:
            href = link.get_attribute("href") or ""
            if (href.startswith("http") and
                    "pinterest.com" not in href and
                    "facebook.com" not in href and
                    "instagram.com" not in href and
                    "twitter.com" not in href and
                    "tiktok.com" not in href and
                    "youtube.com" not in href and
                    "google.com" not in href and
                    re.match(r"https?://[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}", href)):
                return href.strip()

    except Exception as e:
        print(f"        [!] Profile website error: {e}")

    return None


# ─────────────────────────────────────────────
#  CHECK WEBSITE TRAFFIC VIA AHREFS
# ─────────────────────────────────────────────
def get_website_traffic(driver, website_url):
    """Check website traffic using siteworthtraffic.com (no Cloudflare block)."""
    try:
        domain = re.sub(r"https?://(www\.)?", "", website_url).split("/")[0].strip()
        if not domain:
            return None

        print(f"        Checking traffic: {domain}")
        try:
            driver.get(f"https://ahrefs.com/traffic-checker/?target={domain}&mode=subdomains")
        except Exception:
            return None
        time.sleep(AHREFS_WAIT + 3)  # Give Ahrefs extra time to load

        # Try Ahrefs-specific selectors
        for sel in [
            "[data-tf='organic-traffic']",
            ".traffic-value",
            "[class*='organicTraffic']",
            "[class*='trafficValue']",
            "p[class*='value']",
            "span[class*='value']",
            "[class*='metric'] [class*='value']",
        ]:
            try:
                els = driver.find_elements(By.CSS_SELECTOR, sel)
                for el in els:
                    val = parse_number(el.text.strip())
                    if val and val > 0:
                        return val
            except Exception:
                continue

        # Fallback: scan full page text for traffic patterns
        try:
            body = driver.find_element(By.TAG_NAME, "body").text
            # Check if blocked by Cloudflare
            if "cloudflare" in body.lower() or "captcha" in body.lower() or "access denied" in body.lower():
                print("        [!] Ahrefs blocked — try opening Ahrefs in Firefox first")
                return None
            for pattern in [
                r"Organic\s+traffic[^\d]*(\d[\d,\.]*[KMB]?)",
                r"(\d[\d,\.]*[KMB]?)\s*organic\s*(?:visitors|traffic|visits)",
                r"Traffic[^\d]*(\d[\d,\.]*[KMB]?)",
                r"(\d[\d,\.]*[KMB]?)\s*visitors",
            ]:
                m = re.search(pattern, body, re.IGNORECASE)
                if m:
                    val = parse_number(m.group(1))
                    if val is not None:
                        return val
        except Exception:
            pass

    except Exception as e:
        print(f"        [!] Traffic check error: {e}")

    return None


# ─────────────────────────────────────────────
#  SCRAPE PINS + CHECK PROFILES + WEBSITE + TRAFFIC
# ─────────────────────────────────────────────
def scrape_pins(driver, keyword, main_niche, sub_niche=None):
    label = f"SUB: {sub_niche}" if sub_niche else "MAIN NICHE"
    print(f"\n{'─'*55}")
    print(f"  Searching: '{keyword}'  [{label}]")
    print(f"{'─'*55}")

    results     = []
    seen_profiles = set()

    try:
        search_url = f"https://www.pinterest.com/search/pins/?q={keyword.replace(' ', '%20')}&rs=typed"
        driver.get(search_url)
        time.sleep(PAGE_LOAD_WAIT)
        dismiss_popup(driver)

        # ── Collect first 20 pin URLs ──
        pins     = []
        attempts = 0
        while len(pins) < MAX_PINS_PER_SEARCH and attempts < 6:
            links = driver.find_elements(By.CSS_SELECTOR, "a[href*='/pin/']")
            for lnk in links:
                href = lnk.get_attribute("href")
                if href and "/pin/" in href and href not in [p["url"] for p in pins]:
                    pins.append({"url": href})
                if len(pins) >= MAX_PINS_PER_SEARCH:
                    break
            if len(pins) < MAX_PINS_PER_SEARCH:
                driver.execute_script("window.scrollBy(0, 1500);")
                time.sleep(SCROLL_PAUSE)
                attempts += 1

        pins = pins[:MAX_PINS_PER_SEARCH]
        print(f"[+] Collected {len(pins)} pins — checking profiles...")

        for i, pin in enumerate(pins):
            print(f"\n  Pin {i+1}/{len(pins)}")
            try:
                try:
                    driver.get(pin["url"])
                except (TimeoutException, WebDriverException):
                    print(f"    [!] Page load timeout — skipping pin {i+1}")
                    continue
                time.sleep(3)
                dismiss_popup(driver)

                # Find profile URL from pin
                profile_url  = None
                profile_name = None

                for sel in [
                    "a[data-test-id='creator-profile-link']",
                    "[data-test-id='pin-closeup-user'] a",
                    "a[data-test-id='user-rep-link']",
                ]:
                    try:
                        el = driver.find_element(By.CSS_SELECTOR, sel)
                        profile_url  = el.get_attribute("href")
                        profile_name = el.text.strip()
                        if profile_url:
                            break
                    except Exception:
                        continue

                # Fallback profile detection
                if not profile_url:
                    for lnk in driver.find_elements(By.CSS_SELECTOR, "a[href]"):
                        href = lnk.get_attribute("href") or ""
                        parts = href.rstrip("/").split("/")
                        if (href.startswith("https://www.pinterest.com/") and
                                "/pin/" not in href and len(parts) == 5):
                            profile_url  = href
                            profile_name = parts[-1]
                            break

                if not profile_url or profile_url in seen_profiles:
                    print("    Skipped (no profile or duplicate)")
                    continue

                seen_profiles.add(profile_url)
                clean_profile = profile_url.split("?")[0].rstrip("/")

                # ── Visit profile — check monthly views ──
                try:
                    driver.get(clean_profile)
                except (TimeoutException, WebDriverException):
                    print(f"    [!] Profile load timeout — skipping")
                    continue
                time.sleep(3)
                dismiss_popup(driver)

                view_text = None
                for sel in [
                    "[data-test-id='profile-monthly-views']",
                    "[class*='monthlyViews']",
                    "[class*='monthly']",
                ]:
                    try:
                        el = driver.find_element(By.CSS_SELECTOR, sel)
                        view_text = el.text.strip()
                        if view_text:
                            break
                    except Exception:
                        continue

                if not view_text:
                    try:
                        body = driver.find_element(By.TAG_NAME, "body").text
                        m = re.search(
                            r"([\d,.]+\s*[KMBkmb]?)\s*(?:monthly\s*views|Monthly\s*Views)",
                            body
                        )
                        if m:
                            view_text = m.group(1).strip()
                    except Exception:
                        pass

                view_int = parse_number(view_text)

                if not profile_name:
                    try:
                        profile_name = driver.find_element(
                            By.CSS_SELECTOR, "h1, [data-test-id='profile-name']"
                        ).text.strip()
                    except Exception:
                        profile_name = clean_profile.rstrip("/").split("/")[-1]

                print(f"    Profile : {profile_name}")
                print(f"    Views   : {fmt(view_int)}")

                # ── FILTER 1: Pinterest views check ──
                min_views_raw = int(MIN_PINTEREST_VIEWS_M * 1_000_000)
                if not view_int or view_int < min_views_raw:
                    print(f"    ✗ Below {min_views_raw:,} views — skip")
                    continue

                print(f"    ✓ Pinterest filter passed ({fmt(view_int)})")

                # ── Get website URL ──
                website_url = get_website_from_profile(driver, clean_profile)
                print(f"    Website : {website_url or 'Not found'}")

                if not website_url:
                    print("    ✗ No website — skip")
                    continue

                # ── FILTER 2: Website traffic 200+ ──
                traffic = get_website_traffic(driver, website_url)
                print(f"    Traffic : {fmt(traffic)}")

                if traffic is None:
                    traffic_display = "Unknown"
                    traffic_int     = 0
                    print("    ~ Traffic unknown — saving anyway")
                elif traffic < MIN_WEBSITE_TRAFFIC:
                    print(f"    ✗ Below {MIN_WEBSITE_TRAFFIC} visitors — skip")
                    continue
                else:
                    traffic_display = fmt(traffic)
                    traffic_int     = traffic
                    print(f"    ✓ Traffic filter passed ({fmt(traffic)})")

                result = {
                    "profile_name"       : profile_name,
                    "profile_url"        : clean_profile,
                    "pinterest_views"    : fmt(view_int),
                    "pinterest_views_int": view_int,
                    "website_url"        : website_url,
                    "website_traffic"    : traffic_display,
                    "website_traffic_int": traffic_int,
                    "main_niche"         : main_niche,
                    "sub_niche"          : sub_niche if sub_niche else "",
                }
                results.append(result)

            except Exception as e:
                print(f"    [!] Error on pin {i+1}: {e}")
                continue

    except Exception as e:
        print(f"[!] Search error: {e}")
        # If Pinterest blocked — note it
        if "search server" in str(e).lower() or "could not connect" in str(e).lower():
            print(f"[!] Pinterest blocked this proxy — try restarting bot to get new proxy")

    print(f"\n[+] Qualified from '{keyword}': {len(results)}")
    return results


# ─────────────────────────────────────────────
#  SAVE TO EXCEL
# ─────────────────────────────────────────────
def save_to_excel(results, main_niche):
    wb  = openpyxl.Workbook()
    ws  = wb.active
    ws.title = "Pinterest Research"

    RED       = "E60023"
    WHITE     = "FFFFFF"
    PINK      = "FFF0F0"
    GRAY      = "F5F5F5"
    BLUE      = "0563C1"

    hdr_font  = Font(name="Calibri", bold=True, color=WHITE, size=11)
    hdr_fill  = PatternFill(start_color=RED,  end_color=RED,  fill_type="solid")
    main_fill = PatternFill(start_color=PINK, end_color=PINK, fill_type="solid")
    sub_fill  = PatternFill(start_color=GRAY, end_color=GRAY, fill_type="solid")
    center    = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left      = Alignment(horizontal="left",   vertical="center", wrap_text=True)
    thin      = Side(style="thin", color="DDDDDD")
    border    = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Title
    ws.merge_cells("A1:G1")
    ws["A1"].value     = f"Pinterest Research  ·  {main_niche.title()}  ·  {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    ws["A1"].font      = Font(name="Calibri", bold=True, size=13, color=RED)
    ws["A1"].alignment = center
    ws.row_dimensions[1].height = 28

    # Summary
    ws.merge_cells("A2:G2")
    ws["A2"].value     = (f"Filters: Pinterest ≥ {MIN_PINTEREST_VIEWS_M}M views  |  "
                          f"Website ≥ {MIN_WEBSITE_TRAFFIC} visitors/mo  |  "
                          f"Total: {len(results)} profiles")
    ws["A2"].font      = Font(name="Calibri", italic=True, size=10, color="666666")
    ws["A2"].alignment = center
    ws.row_dimensions[2].height = 18

    # Headers
    headers    = ["Profile Name","Profile URL","Pinterest Views","Website URL","Website Traffic","Main Niche","Sub Niche"]
    col_widths = [24, 42, 18, 42, 18, 18, 28]
    for col, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell           = ws.cell(row=3, column=col, value=h)
        cell.font      = hdr_font
        cell.fill      = hdr_fill
        cell.alignment = center
        cell.border    = border
        ws.column_dimensions[cell.column_letter].width = w
    ws.row_dimensions[3].height = 22

    # Rows
    for ri, r in enumerate(results, start=4):
        fill     = main_fill if r["sub_niche"] == "" else sub_fill
        row_data = [
            r["profile_name"],
            r["profile_url"],
            r["pinterest_views"],
            r["website_url"],
            r["website_traffic"],
            r["main_niche"].title(),
            r["sub_niche"].title() if r["sub_niche"] else "— Main —",
        ]
        for col, val in enumerate(row_data, 1):
            cell           = ws.cell(row=ri, column=col, value=val)
            cell.font      = Font(name="Calibri", size=11)
            cell.fill      = fill
            cell.border    = border
            cell.alignment = center if col in (3, 5) else left
            if col in (2, 4) and val and str(val).startswith("http"):
                cell.hyperlink = val
                cell.font      = Font(name="Calibri", size=11, color=BLUE, underline="single")
        ws.row_dimensions[ri].height = 18

    ws.freeze_panes    = "A4"
    ws.auto_filter.ref = f"A3:G{2+len(results)}"

    # One single Excel file per session — named by niche only
    filename = f"pinterest_{main_niche.replace(' ','_')}_results.xlsx"
    desktop  = os.path.join(os.path.expanduser("~"), "Desktop")
    path     = os.path.join(desktop if os.path.exists(desktop) else os.getcwd(), filename)
    wb.save(path)
    print(f"\n✅ Excel saved → {path}")
    return path


# ─────────────────────────────────────────────
#  MAIN
# ─────────────────────────────────────────────
def get_keywords_from_excel(file_path):
    """Read keywords from Column A of an Excel file."""
    try:
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active
        keywords = []
        for row in ws.iter_rows(min_col=1, max_col=1, values_only=True):
            val = row[0]
            if val and str(val).strip():
                keywords.append(str(val).strip().lower())
        # Remove header if it looks like a header
        if keywords and keywords[0] in ["keyword", "keywords", "niche", "niches", "query"]:
            keywords = keywords[1:]
        print(f"[+] Loaded {len(keywords)} keywords from file")
        return keywords
    except Exception as e:
        print(f"[!] Error reading Excel file: {e}")
        return []


def run_bot():
    global MIN_PINTEREST_VIEWS_M, MIN_WEBSITE_TRAFFIC

    print("=" * 60)
    print("   Pinterest Niche Research Bot  v8 (Firefox)")
    print("=" * 60)

    # ── Ask for custom criteria ──
    print("\n--- CRITERIA SETTINGS ---")
    print("  Enter views as exact number: 500000 = 500K, 1000000 = 1M, 2000000 = 2M")
    views_input = input(f"Minimum Pinterest profile views (default {MIN_PINTEREST_VIEWS_M * 1000000:.0f}): ").strip()
    if views_input:
        try:
            # Accept raw numbers: 500000, 1000000, 2000000
            # Also accept K/M shortcuts: 500k, 1m, 2m
            clean = views_input.upper().strip()
            if 'M' in clean:
                MIN_PINTEREST_VIEWS_M = float(clean.replace('M','').strip())
            elif 'K' in clean:
                MIN_PINTEREST_VIEWS_M = float(clean.replace('K','').strip()) / 1000
            else:
                MIN_PINTEREST_VIEWS_M = int(clean) / 1_000_000
            print(f"[+] Min views set to: {int(MIN_PINTEREST_VIEWS_M * 1_000_000):,}")
        except Exception:
            print(f"[!] Invalid input — using default {int(MIN_PINTEREST_VIEWS_M * 1_000_000):,}")

    traffic_input = input(f"Minimum website visitors/mo (default {MIN_WEBSITE_TRAFFIC}): ").strip()
    if traffic_input:
        try:
            MIN_WEBSITE_TRAFFIC = int(traffic_input)
        except Exception:
            print(f"[!] Invalid input — using default {MIN_WEBSITE_TRAFFIC}")

    print(f"\n[+] Criteria set: Pinterest ≥ {MIN_PINTEREST_VIEWS_M}M views | Website ≥ {MIN_WEBSITE_TRAFFIC} visitors/mo")

    # ── Ask for mode ──
    print("\n--- SELECT MODE ---")
    print("  1 = Type a niche manually (bot finds sub-niches automatically)")
    print("  2 = Load keywords from Excel file (no sub-niche search)")
    mode = input("\nEnter mode (1 or 2): ").strip()

    keywords    = []
    main_niche  = "research"

    if mode == "2":
        file_path = input("Enter full path to your Excel file\n(e.g. C:\\Users\\Zarafshan\\Desktop\\keywords.xlsx): ").strip().strip('"')
        if not os.path.exists(file_path):
            print(f"[!] File not found: {file_path}")
            input("Press Enter to exit...")
            return
        keywords   = get_keywords_from_excel(file_path)
        main_niche = os.path.splitext(os.path.basename(file_path))[0]
        if not keywords:
            print("[!] No keywords found in file. Exiting.")
            input("Press Enter to exit...")
            return
        print(f"[+] Keywords to check: {keywords}")
    else:
        niche = input("\nEnter your niche (e.g. home decor): ").strip().lower()
        if not niche:
            print("[!] No niche entered. Exiting.")
            return
        keywords   = [niche]
        main_niche = niche

    driver      = setup_driver()
    all_results = []

    try:
        open_pinterest(driver)

        if mode == "2":
            # ── MODE 2: Keywords from file, no sub-niches ──
            print("\n" + "═"*55)
            print("  MODE 2 — Keyword File Search")
            print("═"*55)
            for i, kw in enumerate(keywords):
                print(f"\n[Keyword {i+1}/{len(keywords)}] {kw}")
                results = scrape_pins(driver, kw, kw, sub_niche=None)
                all_results.extend(results)
                # Auto-save after each keyword
                if all_results:
                    save_to_excel(all_results, main_niche)
                    print(f"    💾 Auto-saved ({len(all_results)} profiles so far)")
                time.sleep(2)
        else:
            # ── MODE 1: Manual niche + sub-niches ──
            niche = keywords[0]

            print("\n" + "═"*55)
            print("  PHASE 1 — Main Niche")
            print("═"*55)
            results = scrape_pins(driver, niche, niche)
            all_results.extend(results)
            if all_results:
                save_to_excel(all_results, main_niche)
                print(f"    💾 Auto-saved ({len(all_results)} profiles so far)")

            print("\n" + "═"*55)
            print("  PHASE 2 — Sub-Niche Discovery")
            print("═"*55)
            sub_niches = get_sub_niches(driver, niche)

            print("\n" + "═"*55)
            print("  PHASE 3 — Sub-Niche Searches")
            print("═"*55)
            for sub in sub_niches:
                results = scrape_pins(driver, sub, niche, sub_niche=sub)
                all_results.extend(results)
                if all_results:
                    save_to_excel(all_results, main_niche)
                    print(f"    💾 Auto-saved ({len(all_results)} profiles so far)")
                time.sleep(2)

    except KeyboardInterrupt:
        print("\n[!] Stopped by user.")
    except Exception as e:
        print(f"\n[!] Unexpected error: {e}")
    finally:
        driver.quit()
        print("\n[*] Browser closed.")

    # Deduplicate
    seen, unique = set(), []
    for r in all_results:
        key = r.get("website_url", "").lower().rstrip("/")
        if key and key not in seen:
            seen.add(key)
            unique.append(r)

    unique.sort(key=lambda x: (int(x["sub_niche"] != ""), -x["pinterest_views_int"]))

    print("\n" + "═"*55)
    print("  FINAL SUMMARY")
    print("═"*55)
    print(f"  Total qualifying profiles : {len(unique)}")

    if unique:
        save_to_excel(unique, main_niche)
        print("\n✅ Done! Check your Desktop for the Excel file.")
    else:
        print("\n[!] No qualifying profiles found.")

    input("\nPress Enter to exit...")


if __name__ == "__main__":
    run_bot()
