"""
PinSearchPro — Flask API Server (Firefox Headless)
====================================================
Uses headless Firefox + Selenium.
No login, no cookies — opens Pinterest explore page directly.
"""

import os, re, time, uuid, threading
from datetime import datetime
from flask import Flask, request, jsonify, send_file
from flask_cors import CORS
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from selenium import webdriver
from selenium.webdriver.firefox.options import Options as FirefoxOptions
from selenium.webdriver.firefox.service import Service as FirefoxService
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException, WebDriverException
from selenium.webdriver.common.keys import Keys

app = Flask(__name__)
CORS(app)

jobs = {}
OUTPUT_DIR = "/tmp/pinsearch_results"
os.makedirs(OUTPUT_DIR, exist_ok=True)

SCROLL_PAUSE   = 2.0
PAGE_LOAD_WAIT = 5


# ─────────────────────────────────────────────
#  FIREFOX DRIVER SETUP
# ─────────────────────────────────────────────
def setup_driver():
    options = FirefoxOptions()
    options.add_argument("--headless")
    options.add_argument("--width=1920")
    options.add_argument("--height=1080")
    options.set_preference("dom.webdriver.enabled", False)
    options.set_preference("useAutomationExtension", False)
    options.set_preference("general.useragent.override",
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:120.0) Gecko/20100101 Firefox/120.0")
    options.set_preference("permissions.default.image", 2)  # block images = faster
    options.set_preference("javascript.enabled", True)

    # Find geckodriver
    import shutil
    gecko = shutil.which("geckodriver") or "/usr/local/bin/geckodriver"
    print(f"[+] geckodriver: {gecko}")

    # Find Firefox
    firefox_paths = [
        "/usr/bin/firefox-esr",
        "/usr/bin/firefox",
        shutil.which("firefox-esr") or "",
        shutil.which("firefox") or "",
    ]
    for p in firefox_paths:
        if p and os.path.exists(p):
            options.binary_location = p
            print(f"[+] Firefox: {p}")
            break

    service = FirefoxService(executable_path=gecko)
    driver = webdriver.Firefox(service=service, options=options)
    driver.set_page_load_timeout(30)
    driver.set_script_timeout(30)
    return driver


# ─────────────────────────────────────────────
#  HELPERS
# ─────────────────────────────────────────────
def update_progress(job_id, msg):
    if job_id in jobs:
        jobs[job_id]["progress"] = msg
        print(f"[{job_id[:8]}] {msg}")

def parse_number(text):
    if not text: return None
    t = str(text).strip().upper().replace(",","").replace("+","").replace(" ","")
    try:
        if "B" in t: return int(float(t.replace("B",""))*1_000_000_000)
        if "M" in t: return int(float(t.replace("M",""))*1_000_000)
        if "K" in t: return int(float(t.replace("K",""))*1_000)
        return int(float(re.sub(r"[^\d.]","",t)))
    except: return None

def fmt(n):
    if n is None: return "N/A"
    if n >= 1_000_000: return f"{n/1_000_000:.1f}M"
    if n >= 1_000: return f"{n/1_000:.1f}K"
    return str(n)

def dismiss_popups(driver):
    """Close any login/signup popups."""
    selectors = [
        "button[data-test-id='closeup-close-button']",
        "button[aria-label='Close']",
        "button[aria-label='close']",
        "div[data-test-id='interstitial-close-button']",
        "button[data-test-id='simple-dialog-close-button']",
        "[class*='closeButton']",
    ]
    for sel in selectors:
        try:
            btn = driver.find_element(By.CSS_SELECTOR, sel)
            btn.click()
            time.sleep(0.5)
            return True
        except: continue

    # Try Escape key
    try:
        driver.find_element(By.TAG_NAME, "body").send_keys(Keys.ESCAPE)
        time.sleep(0.5)
    except: pass
    return False


# ─────────────────────────────────────────────
#  GET WEBSITE FROM PROFILE
# ─────────────────────────────────────────────
def get_website_from_profile(driver, profile_url):
    try:
        driver.get(profile_url)
        time.sleep(PAGE_LOAD_WAIT)
        dismiss_popups(driver)

        for sel in [
            "a[data-test-id='profile-website']",
            "a[data-test-id='user-website-url']",
            "[data-test-id='profile-website-link'] a",
            "[class*='websiteUrl'] a",
            "[class*='website'] a",
        ]:
            try:
                el = driver.find_element(By.CSS_SELECTOR, sel)
                href = el.get_attribute("href")
                if href and "pinterest.com" not in href and href.startswith("http"):
                    return href.strip()
            except: continue

        # Fallback: scan external links
        skip = ["pinterest.com","facebook.com","instagram.com",
                "twitter.com","tiktok.com","youtube.com","google.com"]
        for link in driver.find_elements(By.CSS_SELECTOR, "a[href]"):
            href = link.get_attribute("href") or ""
            if (href.startswith("http") and
                    not any(d in href for d in skip) and
                    re.match(r"https?://[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}", href)):
                return href.strip()
    except: pass
    return None


# ─────────────────────────────────────────────
#  GET WEBSITE TRAFFIC
# ─────────────────────────────────────────────
def get_website_traffic(driver, website_url):
    try:
        domain = re.sub(r"https?://(www\.)?","", website_url).split("/")[0].strip()
        if not domain: return None

        driver.get(f"https://ahrefs.com/traffic-checker/?target={domain}&mode=subdomains")
        time.sleep(8)

        for sel in [
            "[data-tf='organic-traffic']",
            ".traffic-value",
            "[class*='organicTraffic']",
            "[class*='trafficValue']",
        ]:
            try:
                els = driver.find_elements(By.CSS_SELECTOR, sel)
                for el in els:
                    val = parse_number(el.text.strip())
                    if val and val > 0: return val
            except: continue

        # Fallback: scan page text
        try:
            body = driver.find_element(By.TAG_NAME, "body").text
            for pattern in [
                r"Organic\s+traffic[^\d]*(\d[\d,\.]*[KMB]?)",
                r"(\d[\d,\.]*[KMB]?)\s*organic\s*(?:visitors|traffic)",
                r"Traffic[^\d]*(\d[\d,\.]*[KMB]?)",
            ]:
                m = re.search(pattern, body, re.IGNORECASE)
                if m:
                    val = parse_number(m.group(1))
                    if val: return val
        except: pass
    except: pass
    return None


# ─────────────────────────────────────────────
#  SCRAPE PINTEREST
# ─────────────────────────────────────────────
def scrape_pinterest(driver, keyword, min_views, min_traffic, max_pins, job_id):
    results = []
    seen_profiles = set()

    try:
        # Go directly to search — works without login on Firefox
        search_url = f"https://www.pinterest.com/search/pins/?q={keyword.replace(' ','%20')}&rs=typed"
        update_progress(job_id, f"Opening Pinterest search for '{keyword}'...")
        driver.get(search_url)
        time.sleep(PAGE_LOAD_WAIT)
        dismiss_popups(driver)

        # Check if blocked by login wall
        if "login" in driver.current_url.lower():
            # Try explore page instead
            update_progress(job_id, "Trying explore page...")
            driver.get("https://www.pinterest.com/ideas/")
            time.sleep(PAGE_LOAD_WAIT)
            dismiss_popups(driver)
            # Now search
            driver.get(search_url)
            time.sleep(PAGE_LOAD_WAIT)
            dismiss_popups(driver)

        # Collect pin URLs by scrolling
        update_progress(job_id, "Scrolling to collect pins...")
        pins = []
        attempts = 0
        while len(pins) < max_pins and attempts < 8:
            links = driver.find_elements(By.CSS_SELECTOR, "a[href*='/pin/']")
            for lnk in links:
                href = lnk.get_attribute("href")
                if href and "/pin/" in href and href not in [p["url"] for p in pins]:
                    pins.append({"url": href})
                if len(pins) >= max_pins: break
            if len(pins) < max_pins:
                driver.execute_script("window.scrollBy(0, 2000);")
                time.sleep(SCROLL_PAUSE)
                dismiss_popups(driver)
                attempts += 1

        pins = pins[:max_pins]
        total = len(pins)
        update_progress(job_id, f"Found {total} pins — checking each profile...")

        if total == 0:
            update_progress(job_id, "No pins found — Pinterest may be showing a login wall.")
            return results

        for i, pin in enumerate(pins):
            update_progress(job_id, f"Checking pin {i+1}/{total}...")
            try:
                try:
                    driver.get(pin["url"])
                except (TimeoutException, WebDriverException):
                    continue
                time.sleep(3)
                dismiss_popups(driver)

                # Find profile link
                profile_url = None
                profile_name = None
                for sel in [
                    "a[data-test-id='creator-profile-link']",
                    "[data-test-id='pin-closeup-user'] a",
                    "a[data-test-id='user-rep-link']",
                ]:
                    try:
                        el = driver.find_element(By.CSS_SELECTOR, sel)
                        profile_url  = el.get_attribute("href")
                        profile_name = el.text.strip()
                        if profile_url: break
                    except: continue

                # Fallback profile detection
                if not profile_url:
                    for lnk in driver.find_elements(By.CSS_SELECTOR, "a[href]"):
                        href = lnk.get_attribute("href") or ""
                        parts = href.rstrip("/").split("/")
                        if (href.startswith("https://www.pinterest.com/") and
                                "/pin/" not in href and len(parts) == 5):
                            profile_url  = href
                            profile_name = parts[-1]
                            break

                if not profile_url or profile_url in seen_profiles:
                    continue
                seen_profiles.add(profile_url)
                clean_profile = profile_url.split("?")[0].rstrip("/")

                # Visit profile
                try:
                    driver.get(clean_profile)
                except (TimeoutException, WebDriverException):
                    continue
                time.sleep(3)
                dismiss_popups(driver)

                # Get monthly views
                view_text = None
                for sel in [
                    "[data-test-id='profile-monthly-views']",
                    "[class*='monthlyViews']",
                    "[class*='monthly']",
                ]:
                    try:
                        el = driver.find_element(By.CSS_SELECTOR, sel)
                        view_text = el.text.strip()
                        if view_text: break
                    except: continue

                if not view_text:
                    try:
                        body = driver.find_element(By.TAG_NAME, "body").text
                        m = re.search(
                            r"([\d,.]+\s*[KMBkmb]?)\s*(?:monthly\s*views|Monthly\s*Views)",
                            body)
                        if m: view_text = m.group(1).strip()
                    except: pass

                view_int = parse_number(view_text)

                if not profile_name:
                    try:
                        profile_name = driver.find_element(
                            By.CSS_SELECTOR, "h1, [data-test-id='profile-name']"
                        ).text.strip()
                    except:
                        profile_name = clean_profile.rstrip("/").split("/")[-1]

                update_progress(job_id,
                    f"Pin {i+1}/{total} — {profile_name} — "
                    f"views: {fmt(view_int)} (need {fmt(min_views)})")

                # FILTER 1: views
                if not view_int or view_int < min_views:
                    update_progress(job_id, f"{profile_name} ✗ views too low")
                    continue

                # Get website
                update_progress(job_id, f"{profile_name} ✓ views — getting website...")
                website_url = get_website_from_profile(driver, clean_profile)
                if not website_url:
                    update_progress(job_id, f"{profile_name} ✗ no website")
                    continue

                # FILTER 2: traffic
                traffic = None
                if min_traffic > 0:
                    update_progress(job_id, f"{profile_name} — checking traffic...")
                    traffic = get_website_traffic(driver, website_url)
                    if traffic is not None and traffic < min_traffic:
                        update_progress(job_id, f"{profile_name} ✗ traffic too low ({fmt(traffic)})")
                        continue

                result = {
                    "profile_name":        profile_name,
                    "profile_url":         clean_profile,
                    "pinterest_views":     fmt(view_int),
                    "pinterest_views_int": view_int,
                    "website_url":         website_url,
                    "website_traffic":     fmt(traffic) if traffic else "Unknown",
                    "website_traffic_int": traffic or 0,
                }
                results.append(result)
                jobs[job_id]["results"] = list(results)
                update_progress(job_id,
                    f"✓ {profile_name} qualified! ({len(results)} found so far)")

            except Exception as e:
                update_progress(job_id, f"Pin {i+1} error: {e}")
                continue

    except Exception as e:
        update_progress(job_id, f"Search error: {e}")

    return results


# ─────────────────────────────────────────────
#  EXCEL EXPORT
# ─────────────────────────────────────────────
def save_to_excel(results, keyword, job_id):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Pinterest Research"

    RED="E60023"; WHITE="FFFFFF"; PINK="FFF0F0"; BLUE="0563C1"
    hdr_font = Font(name="Calibri", bold=True, color=WHITE, size=11)
    hdr_fill = PatternFill(start_color=RED, end_color=RED, fill_type="solid")
    row_fill = PatternFill(start_color=PINK, end_color=PINK, fill_type="solid")
    center   = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left     = Alignment(horizontal="left",   vertical="center", wrap_text=True)
    thin     = Side(style="thin", color="DDDDDD")
    border   = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.merge_cells("A1:F1")
    ws["A1"].value     = f"Pinterest Research · {keyword.title()} · {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    ws["A1"].font      = Font(name="Calibri", bold=True, size=13, color=RED)
    ws["A1"].alignment = center
    ws.row_dimensions[1].height = 28

    ws.merge_cells("A2:F2")
    ws["A2"].value     = f"Total qualifying profiles: {len(results)}"
    ws["A2"].font      = Font(name="Calibri", italic=True, size=10, color="666666")
    ws["A2"].alignment = center

    headers    = ["Profile Name","Pinterest URL","Monthly Views","Website","Traffic","Keyword"]
    col_widths = [24, 40, 18, 40, 16, 20]
    for col, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=3, column=col, value=h)
        cell.font=hdr_font; cell.fill=hdr_fill
        cell.alignment=center; cell.border=border
        ws.column_dimensions[cell.column_letter].width = w
    ws.row_dimensions[3].height = 22

    for ri, r in enumerate(results, start=4):
        row_data = [
            r["profile_name"], r["profile_url"],
            r["pinterest_views"], r["website_url"],
            r["website_traffic"], keyword.title(),
        ]
        for col, val in enumerate(row_data, 1):
            cell = ws.cell(row=ri, column=col, value=val)
            cell.font=Font(name="Calibri", size=11)
            cell.fill=row_fill; cell.border=border
            cell.alignment=center if col in (3,5) else left
            if col in (2,4) and val and str(val).startswith("http"):
                cell.hyperlink = val
                cell.font=Font(name="Calibri", size=11, color=BLUE, underline="single")
        ws.row_dimensions[ri].height = 18

    ws.freeze_panes = "A4"
    ws.auto_filter.ref = f"A3:F{2+len(results)}"

    path = os.path.join(OUTPUT_DIR, f"{job_id}.xlsx")
    wb.save(path)
    return path


# ─────────────────────────────────────────────
#  BACKGROUND JOB
# ─────────────────────────────────────────────
def run_search_job(job_id, keyword, min_views, min_traffic, max_pins):
    jobs[job_id]["status"] = "running"
    driver = None
    try:
        update_progress(job_id, "Starting Firefox browser...")
        driver = setup_driver()
        update_progress(job_id, "Firefox ready!")

        results = scrape_pinterest(driver, keyword, min_views, min_traffic, max_pins, job_id)

        # Deduplicate
        seen, unique = set(), []
        for r in results:
            key = r.get("website_url","").lower().rstrip("/")
            if key and key not in seen:
                seen.add(key); unique.append(r)

        unique.sort(key=lambda x: -x["pinterest_views_int"])
        jobs[job_id]["results"] = unique

        if unique:
            path = save_to_excel(unique, keyword, job_id)
            jobs[job_id]["excel_path"] = path
            update_progress(job_id, f"Done! {len(unique)} profiles found. Excel ready.")
        else:
            update_progress(job_id, "Done — no qualifying profiles found. Try lower min views.")

        jobs[job_id]["status"] = "done"

    except Exception as e:
        jobs[job_id]["status"] = "error"
        jobs[job_id]["error"]  = str(e)
        update_progress(job_id, f"Error: {e}")
    finally:
        if driver:
            try: driver.quit()
            except: pass


# ─────────────────────────────────────────────
#  API ROUTES
# ─────────────────────────────────────────────
@app.route("/health", methods=["GET"])
def health():
    return jsonify({"status": "ok", "message": "PinSearchPro API — Firefox headless mode"})

@app.route("/search", methods=["POST"])
def start_search():
    data = request.get_json()
    if not data or not data.get("keyword"):
        return jsonify({"error": "keyword is required"}), 400

    keyword     = data["keyword"].strip().lower()
    min_views   = int(data.get("min_views", 2_000_000))
    min_traffic = int(data.get("min_traffic", 0))
    max_pins    = max(20, min(int(data.get("max_pins", 20)), 200))

    job_id = str(uuid.uuid4())
    jobs[job_id] = {
        "status": "queued", "progress": "Queued...",
        "results": [], "excel_path": None,
        "error": None, "keyword": keyword,
        "created_at": datetime.now().isoformat(),
    }

    threading.Thread(
        target=run_search_job,
        args=(job_id, keyword, min_views, min_traffic, max_pins),
        daemon=True
    ).start()

    return jsonify({"job_id": job_id, "status": "queued"})

@app.route("/status/<job_id>", methods=["GET"])
def get_status(job_id):
    job = jobs.get(job_id)
    if not job: return jsonify({"error": "Job not found"}), 404
    return jsonify({
        "job_id": job_id, "status": job["status"],
        "progress": job["progress"],
        "result_count": len(job["results"]),
        "results": job["results"],
        "has_excel": job["excel_path"] is not None,
        "error": job["error"],
    })

@app.route("/download/<job_id>", methods=["GET"])
def download_excel(job_id):
    job = jobs.get(job_id)
    if not job: return jsonify({"error": "Job not found"}), 404
    if not job.get("excel_path") or not os.path.exists(job["excel_path"]):
        return jsonify({"error": "Excel not ready"}), 404
    return send_file(
        job["excel_path"], as_attachment=True,
        download_name=f"pinterest_{job['keyword'].replace(' ','_')}_results.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))
    app.run(host="0.0.0.0", port=port, debug=False)
